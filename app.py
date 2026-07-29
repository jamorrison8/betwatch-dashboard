#!/usr/bin/env python3
"""
app.py - Betwatch odds dashboard (Streamlit)

Fetch-once architecture:
    - "Fetch Meeting Data" is the ONLY control that calls the Betwatch API.
      It pulls raw pricing for the selected date/track and caches it in
      st.session_state["raw_meeting_data"].
    - Every slider/filter (commission, stake, bookmakers, thresholds,
      liquidity) recomputes/refilters instantly from that cached raw data.
      No network calls happen outside the fetch button handler.

Setup (local):
    pip install betwatch streamlit openpyxl pandas
    export BETWATCH_API_KEY="your-key-here"
    streamlit run app.py

The dashboard opens in your browser at http://localhost:8501

Setup (Streamlit Community Cloud):
    Add two secrets in the app's Settings -> Secrets:
        BETWATCH_API_KEY = "your-key-here"
        APP_PASSWORD = "choose-a-password"
    No code changes needed - the secrets are picked up automatically below.
"""

import os
from datetime import datetime, date
from io import BytesIO
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st

SYDNEY_TZ = ZoneInfo("Australia/Sydney")


def _format_sydney_time(start_time):
    """
    Betwatch's race.start_time may come back as a datetime or an ISO string,
    with or without timezone info (assume UTC if bare, since that's what
    Betwatch stores). Returns e.g. "2:35 PM" (Sydney local, no tz label
    since the person only ever needs Sydney time) or None if unavailable.
    """
    if start_time is None:
        return None
    dt = start_time
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return None
    if not isinstance(dt, datetime):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    dt_syd = dt.astimezone(SYDNEY_TZ)
    return dt_syd.strftime("%I:%M %p").lstrip("0")

# ---------------------------------------------------------------------------
# Deployment helpers: secrets wiring + password gate (no-ops when run locally
# with an env var and no st.secrets configured)
# ---------------------------------------------------------------------------

def _load_secrets_into_env():
    """
    On Streamlit Community Cloud, BETWATCH_API_KEY is set as a "secret", not
    an OS env var. betwatch.connect() only reads the environment, so copy it
    across if present. Locally (env var already set, no secrets.toml) this
    is a harmless no-op.
    """
    try:
        if "BETWATCH_API_KEY" in st.secrets and not os.environ.get("BETWATCH_API_KEY"):
            os.environ["BETWATCH_API_KEY"] = st.secrets["BETWATCH_API_KEY"]
    except Exception:
        pass  # no secrets.toml at all (e.g. plain local run) - fine


def _require_password():
    """
    Simple password gate using an APP_PASSWORD secret. If no APP_PASSWORD is
    configured (e.g. running locally), the gate is skipped entirely - it only
    activates once you set that secret, which you should do before deploying
    anywhere publicly reachable.
    """
    try:
        required = st.secrets.get("APP_PASSWORD")
    except Exception:
        required = None
    if not required:
        return  # no password configured - local/dev use, skip the gate

    if st.session_state.get("_authed"):
        return

    st.title("Betwatch Odds Dashboard")
    entered = st.text_input("Password", type="password")
    if st.button("Enter"):
        if entered == required:
            st.session_state["_authed"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()  # halt the rest of the script until authenticated


# ---------------------------------------------------------------------------
# Pure calculation layer (no streamlit, no network - safe to unit test)
# ---------------------------------------------------------------------------

PAGE_SIZE = 100

# Starting commission rates by meeting location, in PERCENT.
# These are SEED VALUES ONLY for the slider - Betfair AU racing commission is
# set per state and racing code and changes periodically, so verify against
# Betfair's own published rates rather than trusting this table.
COMMISSION_SEED_BY_LOCATION = {
    "NSW": 8.0, "VIC": 8.0, "QLD": 8.0, "SA": 8.0,
    "WA": 8.0, "TAS": 8.0, "NT": 8.0, "ACT": 8.0, "NZL": 8.0,
}
COMMISSION_SEED_DEFAULT = 8.0

# Only these bookmakers are used - everything else is noise to filter out by
# default. Matched case-insensitively against whatever Betwatch's API returns,
# since display names don't always match exactly. Betfair itself isn't in this
# list because it's the exchange (back/lay columns), not a bookmaker row.
MY_BOOKMAKERS = {
    "tab", "bet365", "sportsbet", "pointsbet", "tabtouch",
    "betdeluxe", "amused",  # "Amused" is how Betdeluxe might show if rebranded
    "betr", "ladbrokes", "neds", "picklebet",
    "swiftbet", "betmakers",  # Betmakers group brand
    "betright", "dabble", "crownbet",
}


def fetch_raw_rows(client, date_from, date_to, track=None):
    """
    Single fetch action: page through the Betwatch API and flatten every
    race/runner/bookmaker combination into raw price dicts. No retention/loss
    maths happens here - just extraction of raw prices and metadata.
    Returns (rows, race_summaries, n_races, n_runners). race_summaries covers
    every race regardless of whether it has usable odds rows, so a resulted
    race with no live prices still surfaces its results.
    """
    from betwatch.types import RaceProjection, RacesFilter

    projection = RaceProjection(
        markets=True,        # bookmaker fixed win markets
        place_markets=False,
        flucs=True,          # opening price is derived from flucs[0]
        links=False,
        betfair=True,
    )

    races = []
    offset = 0
    while True:
        page = client.get_races(projection, RacesFilter(
            date_from=date_from, date_to=date_to,
            tracks=[track] if track else None,
            limit=PAGE_SIZE, offset=offset,
        ))
        if not page:
            break
        races.extend(page)
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE

    rows = []
    race_summaries = []
    n_runners = 0
    for race in races:
        try:
            meeting = getattr(race, "meeting", None)
            track_name = str(getattr(meeting, "track", "Unknown") or "Unknown") if meeting else "Unknown"
            location = str(getattr(meeting, "location", "") or "") if meeting else ""
            race_no = getattr(race, "number", None)
            status = _status_str(race)
            results_str = _resolve_top4_results(race)
            race_time = _format_sydney_time(getattr(race, "start_time", None))

            # Record every race that comes back, independent of whether any
            # runner has usable odds - a resulted race with no live prices
            # must still show its results.
            race_summaries.append({
                "track": track_name,
                "race_no": race_no,
                "status": status,
                "results_str": results_str,
                "race_time": race_time,
            })

            for runner in (getattr(race, "runners", None) or []):
                if getattr(runner, "scratched_time", None):
                    continue
                n_runners += 1

                bf = _betfair_win_market(runner)
                lay_price = lay_size = back_price = None
                total_matched = 0
                if bf is not None:
                    lay_list = getattr(bf, "lay", None) or []
                    back_list = getattr(bf, "back", None) or []
                    lay_price = getattr(lay_list[0], "price", None) if lay_list else None
                    lay_size = getattr(lay_list[0], "size", None) if lay_list else None
                    back_price = getattr(back_list[0], "price", None) if back_list else None
                    total_matched = (getattr(bf, "market_total_matched", None)
                                     or getattr(bf, "total_matched", None) or 0)
                    if lay_price is not None and lay_price <= 1:
                        lay_price = None  # invalid price treated same as missing

                for bm in (getattr(runner, "bookmaker_markets", None) or []):
                    fixed_win = getattr(bm, "fixed_win", None)
                    if not fixed_win:
                        continue
                    price = getattr(fixed_win, "price", None)
                    if not price or price <= 1:
                        continue
                    flucs = getattr(fixed_win, "flucs", None) or []
                    opening = getattr(flucs[0], "price", None) if flucs else None

                    rows.append({
                        "track": track_name,
                        "location": location,
                        "race_no": race_no,
                        "status": status,
                        "runner_no": getattr(runner, "number", None),
                        "runner_name": getattr(runner, "name", "?"),
                        "bookmaker": str(getattr(bm, "bookmaker", "Unknown") or "Unknown"),
                        "fixed_win_price": float(price),
                        "opening_price": opening,
                        "betfair_lay_price": float(lay_price) if lay_price is not None else None,
                        "betfair_lay_size": lay_size,
                        "betfair_back_price": back_price,
                        "total_matched": float(total_matched),
                        "results_str": results_str,
                        "race_time": race_time,
                    })
        except Exception:
            continue  # one malformed race must not kill the fetch

    seen = set()
    deduped = []
    for r in rows:
        key = (r["track"], r["race_no"], r["runner_no"], r["bookmaker"])
        if key not in seen:
            seen.add(key)
            deduped.append(r)
    rows = deduped

    race_summaries.sort(key=lambda x: (x["race_no"] is None, x["race_no"]))

    return rows, race_summaries, len(races), n_runners


def _resolve_top4_results(race):
    """
    Betwatch's `results` field is a list of positions, each a list of runner
    numbers (more than one number in a position = dead heat). Only the top 4
    positions are populated. Resolve runner numbers to names for readability.
    Returns None if the race hasn't been resulted yet.
    """
    raw_results = getattr(race, "results", None)
    if not raw_results:
        return None

    name_by_number = {}
    for runner in (getattr(race, "runners", None) or []):
        num = getattr(runner, "number", None)
        if num is not None:
            name_by_number[num] = getattr(runner, "name", f"#{num}")

    positions = []
    for pos in raw_results[:4]:
        names = [name_by_number.get(n, f"#{n}") for n in (pos or [])]
        if names:
            positions.append("/".join(names))  # "/" joins dead-heated runners

    if not positions:
        return None
    return ",".join(f"{i+1}:{name}" for i, name in enumerate(positions))


def _betfair_win_market(runner):
    """The runner's Betfair WIN market only - never the place market."""
    getter = getattr(runner, "get_betfair_win_market", None)
    if callable(getter):
        return getter()
    for m in (getattr(runner, "betfair_markets", None) or []):
        if getattr(m, "market_name", None) == "win":
            return m
    return None


def _status_str(race):
    s = getattr(race, "status", None)
    if s is None:
        return ""
    return getattr(s, "value", None) or str(s)


def retention_pct(B, L, c, mode):
    """Bonus-bet retention %. snr = stake not returned, sr = stake returned."""
    if B is None or L is None or L - c <= 0:
        return None
    if mode == "snr":
        return (B - 1) * (1 - c) / (L - c) * 100
    return B * (1 - c) / (L - c) * 100  # sr


def back_lay_outcome(B, L, c, stake, mode="cash"):
    """
    Standard back+lay pair maths. Returns (lay_stake, guaranteed_outcome),
    where guaranteed_outcome = min(outcome if back wins, outcome if lay wins).

    mode="cash": real-money qualifying/mug bet. Outcome is a small negative
                 number (the guaranteed loss).
    mode="snr":  stake-not-returned free bet. Only winnings (B-1) are live,
                 no cash stake at risk. Outcome is the positive cash retained.
    mode="sr":   stake-returned free bet. Bet behaves like cash on the back
                 side, but the stake itself costs nothing. Outcome is the
                 positive cash retained (same lay-sizing formula as "cash").
    """
    if B is None or L is None or L - c <= 0:
        return None, None
    if mode == "snr":
        lay_stake = stake * (B - 1) / (L - c)
        outcome_back = stake * (B - 1) - lay_stake * (L - 1)
        outcome_lay = lay_stake * (1 - c)
    elif mode == "sr":
        lay_stake = stake * B / (L - c)
        outcome_back = stake * B - lay_stake * (L - 1)
        outcome_lay = lay_stake * (1 - c)
    else:  # cash
        lay_stake = stake * B / (L - c)
        outcome_back = stake * (B - 1) - lay_stake * (L - 1)
        outcome_lay = lay_stake * (1 - c) - stake
    return lay_stake, min(outcome_back, outcome_lay)


def build_odds_bundle(raw_rows, race_summaries, meta):
    """
    Dense, compact text dump of the cached raw meeting data - built for pasting
    into an LLM chat, not for human readability. Prints EVERY race that came
    back from the fetch (race_summaries), each with its results header, then
    any odds rows for that race underneath. A race with no usable odds rows
    still gets its results header - results must never be dropped just
    because a market has no live prices.
    """
    if not race_summaries:
        return "No data fetched yet."

    lines = [
        f"MEETING_BUNDLE track={meta.get('track','ALL')} "
        f"date_from={meta.get('date_from','?')} date_to={meta.get('date_to','?')} "
        f"fetched={meta.get('when','?')} races={meta.get('races','?')} runners={meta.get('runners','?')}"
    ]

    # Group odds rows by (track, race_no) for quick lookup under each race header
    rows_by_race = {}
    for r in raw_rows:
        key = (r["track"], r["race_no"])
        rows_by_race.setdefault(key, []).append(r)

    for race in race_summaries:
        key = (race["track"], race["race_no"])
        results = race.get("results_str")
        lines.append(
            f"--RACE track={race['track']} no={race['race_no']} status={race['status']} "
            f"results={results if results else 'not yet resulted'}"
        )
        for r in rows_by_race.get(key, []):
            lines.append(
                f"R{r['race_no']}|{r['runner_no']}|{r['runner_name']}|"
                f"BOOK:{r['bookmaker']}:{r['fixed_win_price']}(open:{r.get('opening_price')})|"
                f"BFback:{r.get('betfair_back_price')}|BFlay:{r['betfair_lay_price']}|"
                f"BFlaySize:{r.get('betfair_lay_size')}|BFmatched:{r['total_matched']}"
            )
        if key not in rows_by_race:
            lines.append("  (no usable odds rows for this race)")

    return "\n".join(lines)


def _tier_colors(value, thresholds_desc):
    """
    thresholds_desc: list of (cutoff, bg_hex, fg) checked in order, first match
    wins. Used both for "higher is better" (retention %, descending cutoffs)
    and "lower is better" (loss %, ascending cutoffs via negated comparison
    handled by the caller).
    """
    for cutoff, bg, fg in thresholds_desc:
        if value >= cutoff:
            return f"background-color: {bg}; color: {fg}"
    return ""


def style_bonus_table(df):
    """Highlight rows by Retention % - higher is better. Text color switches
    for legibility on darker backgrounds."""
    if df is None or df.empty:
        return df
    tiers = [
        (80, "#0B6623", "white"),   # dark green
        (75, "#8FD18F", "black"),  # light green
        (70, "#F5D742", "black"),  # yellow
        (65, "#F0A030", "black"),  # orange
    ]

    def _row_style(row):
        style = _tier_colors(row["Retention %"], tiers)
        return [style] * len(row)

    float_cols = df.select_dtypes(include="float").columns
    return df.style.apply(_row_style, axis=1).format(precision=2, subset=float_cols)


def style_mug_table(df):
    """Highlight rows by Net % - a positive value is a net profit, negative
    is a net loss. Higher (more profit / smaller loss) is better. Text color
    switches for legibility on darker backgrounds."""
    if df is None or df.empty:
        return df
    tiers = [
        (5, "#0B6623", "white"),    # dark green: net profit >= 5%
        (-10, "#8FD18F", "black"),  # light green: net loss no worse than 10%
        (-15, "#F5D742", "black"),  # yellow: net loss no worse than 15%
        (-20, "#F0A030", "black"),  # orange: net loss no worse than 20%
    ]

    def _row_style(row):
        style = _tier_colors(row["Net %"], tiers)
        return [style] * len(row)

    float_cols = df.select_dtypes(include="float").columns
    return df.style.apply(_row_style, axis=1).format(precision=2, subset=float_cols)



def format_tracker_row(row, kind, race_date_iso, stake, commission_pct, account_prefix=""):
    """
    One tab-separated line matching the bet tracker's expected columns:
    Date | Bookmaker | Type | Category | Race | Runner | Stake | Back odds |
    Lay stake | Lay odds | Commission %
    `row` is a single row (Series) from bonus_view or mug_view.
    account_prefix is prepended to the bookmaker name here only (e.g. "2" for
    a secondary account) - it never touches the main dashboard tables.
    """
    try:
        d = datetime.fromisoformat(race_date_iso)
        date_str = f"{d.strftime('%b')} {d.day}"
    except Exception:
        date_str = race_date_iso or ""

    lay_stake_col = f"Lay Stake (${stake:g})"
    race = f"{row['Track']} R{row['Race #']}"

    fields = [
        date_str,
        f"{account_prefix}{row['Bookmaker']}",
        kind,
        "Racing",
        race,
        str(row["Runner"]),
        f"${stake:,.2f}",
        f"{row['Fixed Win']:g}",
        f"${row.get(lay_stake_col, 0):,.2f}",
        f"{row['BF Lay']:g}",
        f"{commission_pct:.2f}%",
    ]
    return "\t".join(fields)


def compute_tables(raw_rows, commission, stake, free_bet_mode):
    """
    Pure recomputation over the cached raw rows. Called on every rerun;
    MUST NEVER touch the network. Returns (bonus_df, mug_df) unfiltered -
    threshold/bookmaker filtering happens afterwards on the dataframes.
    """
    bonus_records, mug_records = [], []

    for r in raw_rows:
        B, L = r["fixed_win_price"], r["betfair_lay_price"]
        ret = retention_pct(B, L, commission, free_bet_mode)
        if ret is None:
            continue

        lay_size = r["betfair_lay_size"]

        lay_stake_bonus, profit_bonus = back_lay_outcome(B, L, commission, stake, mode=free_bet_mode)
        # Only surface this option if the unmatched liquidity at the lay price
        # actually covers the lay stake required - otherwise it can't be
        # matched at size and shouldn't be shown as a usable opportunity.
        if lay_size is not None and lay_size >= lay_stake_bonus:
            bonus_records.append({
                "Track": r["track"], "Race #": r["race_no"],
                "Race Time": r.get("race_time"),
                "Runner #": r["runner_no"], "Runner": r["runner_name"],
                "Bookmaker": r["bookmaker"],
                "Fixed Win": round(B, 2), "BF Lay": round(L, 2),
                "Lay Size": round(lay_size, 2) if lay_size is not None else None,
                "Retention %": round(ret, 2),
                f"Lay Stake (${stake:g})": round(lay_stake_bonus, 2),
                f"Return (${stake:g} bonus)": round(profit_bonus, 2),
                "Status": r["status"],
            })

        lay_stake, worst = back_lay_outcome(B, L, commission, stake, mode="cash")
        if lay_size is not None and lay_size >= lay_stake:
            net_pct = (worst / stake * 100) if stake else 0.0
            mug_records.append({
                "Track": r["track"], "Race #": r["race_no"],
                "Race Time": r.get("race_time"),
                "Runner #": r["runner_no"], "Runner": r["runner_name"],
                "Bookmaker": r["bookmaker"],
                "Fixed Win": round(B, 2), "BF Lay": round(L, 2),
                "Lay Size": round(lay_size, 2) if lay_size is not None else None,
                "Total Matched": round(r["total_matched"], 2),
                f"Lay Stake (${stake:g})": round(lay_stake, 2),
                f"Loss (${stake:g})": round(-worst, 2),
                "Net %": round(net_pct, 2),
                "Status": r["status"],
        })

    bonus_df = pd.DataFrame(bonus_records)
    mug_df = pd.DataFrame(mug_records)
    if not bonus_df.empty:
        bonus_df = bonus_df.sort_values("Retention %", ascending=False).reset_index(drop=True)
    if not mug_df.empty:
        mug_df = mug_df.sort_values(f"Loss (${stake:g})", ascending=True).reset_index(drop=True)
    return bonus_df, mug_df


def export_view_xlsx(df, sheet_name, assumptions):
    """Write the currently displayed dataframe + an Assumptions sheet to xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")

    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.append(["Assumption", "Value"])
    for k, v in assumptions.items():
        ws_a.append([k, v])
    for cell in ws_a[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws_a.column_dimensions["A"].width = 34
    ws_a.column_dimensions["B"].width = 60

    ws = wb.create_sheet(sheet_name)
    ws.append(["Rank"] + list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        ws.append([i] + [None if pd.isna(v) else v for v in row.tolist()])
    ws.column_dimensions["A"].width = 6
    for idx in range(2, len(df.columns) + 2):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

def seed_commission_from_location(raw_rows):
    """Pick a starting commission % from the most common meeting location."""
    locs = [r["location"] for r in raw_rows if r.get("location")]
    if not locs:
        return COMMISSION_SEED_DEFAULT
    top = max(set(locs), key=locs.count)
    return COMMISSION_SEED_BY_LOCATION.get(top, COMMISSION_SEED_DEFAULT)


def main():
    st.set_page_config(page_title="Betwatch Odds Dashboard", layout="wide")
    _load_secrets_into_env()
    _require_password()  # no-op locally unless APP_PASSWORD secret is set
    st.title("Betwatch Odds Dashboard")

    # ---- Top bar: fetch controls (the ONLY network path) ----
    c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.6, 1.0])
    with c1:
        date_from = st.date_input("Date from", value=date.today())
    with c2:
        date_to = st.date_input("Date to", value=date.today())
    with c3:
        track = st.text_input("Track (optional, blank = all tracks)", value="")
    with c4:
        st.write("")  # vertical spacing
        fetch_clicked = st.button("Fetch Meeting Data", type="primary", width="stretch")

    if fetch_clicked:
        try:
            import betwatch
        except ImportError:
            st.error("betwatch package missing - run: pip install betwatch")
            st.stop()
        try:
            client = betwatch.connect()  # reads BETWATCH_API_KEY from environment
        except Exception as e:
            st.error(f"Could not connect - is BETWATCH_API_KEY set? ({e})")
            st.stop()
        with st.spinner("Fetching from Betwatch..."):
            try:
                rows, race_summaries, n_races, n_runners = fetch_raw_rows(
                    client,
                    date_from.strftime("%Y-%m-%d"),
                    date_to.strftime("%Y-%m-%d"),
                    track.strip() or None,
                )
            except Exception as e:
                st.error(f"API request failed: {e}")
                st.stop()
        st.session_state["raw_meeting_data"] = rows
        st.session_state["race_summaries"] = race_summaries
        st.session_state["fetch_meta"] = {
            "when": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "races": n_races, "runners": n_runners,
            "date_from": date_from.isoformat(), "date_to": date_to.isoformat(),
            "track": track.strip() or "ALL",
        }
        # Seed the commission slider from the fetched meeting locations
        # (set BEFORE the slider widget is instantiated this rerun).
        st.session_state["commission_pct"] = seed_commission_from_location(rows)

    meta = st.session_state.get("fetch_meta")
    if meta:
        st.caption(f"Last fetched {meta['when']} - {meta['races']} races / "
                   f"{meta['runners']} runners ({meta['date_from']} to {meta['date_to']}, "
                   f"track: {meta['track']}). Sliders below recalculate instantly - "
                   f"no re-fetch until you press the button again.")
    else:
        st.info("No data yet - pick a date (and optionally a track), then press "
                "**Fetch Meeting Data**. Everything below works off that one fetch.")

    raw_rows = st.session_state.get("raw_meeting_data", [])
    race_summaries = st.session_state.get("race_summaries", [])

    # ---- Sidebar: instant controls (cached data only, never the network) ----
    with st.sidebar:
        st.header("Controls")
        commission_pct = st.slider(
            "Betfair commission %", 0.0, 15.0,
            st.session_state.get("commission_pct", COMMISSION_SEED_DEFAULT),
            0.5, key="commission_pct",
        )
        st.caption("Seeded per fetch from meeting location as a starting value only - "
                   "AU racing commission varies by state/code and changes over time, "
                   "so verify against Betfair's published rates.")
        stake = st.number_input("Stake $", min_value=1.0, value=50.0, step=5.0)
        free_bet_mode = st.radio(
            "Free bet mode (Bonus tab)", ("snr", "sr"), horizontal=True,
            format_func=lambda m: "Stake not returned (SNR)" if m == "snr" else "Stake returned (SR)",
        )

        all_bookmakers = sorted({r["bookmaker"] for r in raw_rows})
        default_bookmakers = [b for b in all_bookmakers if b.lower() in MY_BOOKMAKERS]
        if not default_bookmakers:
            default_bookmakers = all_bookmakers  # fallback if none matched (e.g. no fetch yet)
        sel_bookmakers = st.multiselect(
            "Bookmakers", all_bookmakers, default=default_bookmakers,
            placeholder="Fetch data to populate",
        )
        if all_bookmakers:
            found_lower = {b.lower() for b in all_bookmakers}
            missing = MY_BOOKMAKERS - found_lower
            # "amused"/"betmakers" are aliases, not expected to match on their own
            missing -= {"amused", "betmakers"}
            if missing:
                st.caption(f"Not seen in this fetch (naming may differ): {', '.join(sorted(missing))}")

        st.divider()
        use_second_account = st.checkbox(
            "Use 2nd account prefix in tracker rows",
            value=False,
            help='Prefixes the bookmaker name with "2" (e.g. "2Sportsbet") in the '
                 "bet tracker copy rows only - does not affect the tables above.",
        )
        account_prefix = "2" if use_second_account else ""

        st.divider()
        target_retention = st.slider("Min retention % (Bonus tab)", 0.0, 120.0, 0.0, 1.0)
        target_loss = st.slider("Max loss $ (Mug tab)", 0.0, 50.0, 50.0, 0.5)
        min_liquidity = st.slider("Min liquidity - BF matched $ (Mug tab)",
                                  0, 100_000, 0, 500)

    commission = commission_pct / 100.0

    # ---- Recompute from cached raw data (every rerun, pure, instant) ----
    bonus_df, mug_df = compute_tables(raw_rows, commission, stake, free_bet_mode)

    loss_col = f"Loss (${stake:g})"
    scanning_all_tracks = (meta.get("track", "ALL") == "ALL") if meta else True

    if not bonus_df.empty:
        bonus_view = bonus_df[
            bonus_df["Bookmaker"].isin(sel_bookmakers)
            & (bonus_df["Retention %"] >= target_retention)
        ]
        if scanning_all_tracks:
            # All-tracks scans can return a huge number of rows - a hard floor
            # keeps the view manageable. A specific-meeting scan shows
            # everything (subject only to the sliders above).
            bonus_view = bonus_view[bonus_view["Retention %"] >= 65]
        bonus_view = bonus_view.reset_index(drop=True)
    else:
        bonus_view = bonus_df

    if not mug_df.empty:
        mug_view = mug_df[
            mug_df["Bookmaker"].isin(sel_bookmakers)
            & (mug_df["Total Matched"] >= min_liquidity)
            & (mug_df[loss_col] <= target_loss)
        ]
        if scanning_all_tracks:
            mug_view = mug_view[mug_view["Net %"] > -20]
        mug_view = mug_view.reset_index(drop=True)
    else:
        mug_view = mug_df

    tab1, tab2, tab3 = st.tabs(["Bonus Conversions", "Mug Bets", "Odds Bundle (copy for Claude)"])

    with tab1:
        st.subheader(f"Bonus conversions - {len(bonus_view)} rows "
                     f"({'SNR' if free_bet_mode == 'snr' else 'SR'}, "
                     f"commission {commission_pct:g}%)")
        if scanning_all_tracks:
            st.caption("Scanning all tracks - hard floor of 65% retention applied to keep this "
                       "manageable. Scan a specific track to see every option, unfiltered.")
        st.dataframe(style_bonus_table(bonus_view), width="stretch", height=520)
        if not bonus_view.empty:
            xlsx = export_view_xlsx(bonus_view, "Bonus Conversions", {
                "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Commission (fraction)": commission,
                "Stake $": stake,
                "Free bet mode": "Stake NOT returned (SNR)" if free_bet_mode == "snr" else "Stake returned (SR)",
                "Formula": ("Retention % = (B - 1) * (1 - c) / (L - c)" if free_bet_mode == "snr"
                            else "Retention % = B * (1 - c) / (L - c)"),
                "Min retention % filter": target_retention,
                "Bookmakers": ", ".join(sel_bookmakers) if sel_bookmakers else "none",
                "Data fetched": meta["when"] if meta else "",
            })
            st.download_button("Export current view to Excel", xlsx,
                               file_name="bonus_conversions.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with st.expander(f"Copy rows for bet tracker ({len(bonus_view)} rows)"):
                kind = "SNR Bonus" if free_bet_mode == "snr" else "SR Bonus"
                race_date_iso = meta.get("date_from") if meta else None
                for _, row in bonus_view.iterrows():
                    label = f"{row['Runner']} - {row['Track']} R{row['Race #']} ({row['Bookmaker']})"
                    st.caption(label)
                    st.code(format_tracker_row(row, kind, race_date_iso, stake, commission_pct, account_prefix),
                             language=None, wrap_lines=False)

    with tab2:
        st.subheader(f"Mug bets - {len(mug_view)} rows "
                     f"(commission {commission_pct:g}%, min liquidity ${min_liquidity:,})")
        if scanning_all_tracks:
            st.caption("Scanning all tracks - hard ceiling of 20% loss applied to keep this "
                       "manageable. Scan a specific track to see every option, unfiltered.")
        st.dataframe(style_mug_table(mug_view), width="stretch", height=520)
        if not mug_view.empty:
            xlsx = export_view_xlsx(mug_view, "Mug Bets", {
                "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Commission (fraction)": commission,
                "Stake $": stake,
                "Min liquidity (BF matched $)": min_liquidity,
                "Max loss $ filter": target_loss,
                "Bookmakers": ", ".join(sel_bookmakers) if sel_bookmakers else "none",
                "Purpose": ("Small real-money back+lay pairs on popular, high-volume "
                            "markets for account sustainability, ranked by smallest loss."),
                "Data fetched": meta["when"] if meta else "",
            })
            st.download_button("Export current view to Excel", xlsx,
                               file_name="mug_bets.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with st.expander(f"Copy rows for bet tracker ({len(mug_view)} rows)"):
                race_date_iso = meta.get("date_from") if meta else None
                for _, row in mug_view.iterrows():
                    label = f"{row['Runner']} - {row['Track']} R{row['Race #']} ({row['Bookmaker']})"
                    st.caption(label)
                    st.code(format_tracker_row(row, "Mug", race_date_iso, stake, commission_pct, account_prefix),
                             language=None, wrap_lines=False)

    with tab3:
        st.subheader("Odds bundle - paste this into a Claude chat")
        st.caption("Dense text dump from the last fetch, filtered to the bookmakers selected "
                   "in the sidebar (results/race headers always show regardless of bookmaker "
                   "filter). Click the copy icon in the top-right of the box below.")
        bundle_rows = [r for r in raw_rows if r["bookmaker"] in sel_bookmakers]
        bundle_text = build_odds_bundle(bundle_rows, race_summaries, meta or {})
        st.code(bundle_text, language=None, wrap_lines=False)


if __name__ == "__main__":
    main()
